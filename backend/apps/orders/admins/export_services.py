import io
from datetime import datetime
from decimal import Decimal
import zoneinfo

from django.utils import timezone
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IST_TZ = zoneinfo.ZoneInfo("Asia/Kolkata")


def format_inr(val):
    num = float(val) if val is not None else 0.0
    return f"Rs. {num:,.2f}"


class NumberedCanvas(canvas.Canvas):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#6b7280"))
        
        # Footer line
        self.setStrokeColor(colors.HexColor("#e5e7eb"))
        self.setLineWidth(0.5)
        self.line(36, 36, 559, 36)

        footer_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(559, 22, footer_text)
        self.drawString(36, 22, "Toy Store - Confidential Admin Sales Report")
        self.restoreState()


class SalesReportPDFGenerator:

    @classmethod
    def generate(cls, report_data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=54,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1e1b4b"),
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#4f46e5"),
        )
        meta_style = ParagraphStyle(
            "ReportMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#4b5563"),
        )
        section_style = ParagraphStyle(
            "ReportSection",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#111827"),
            spaceBefore=12,
            spaceAfter=6,
        )
        cell_header_style = ParagraphStyle(
            "CellHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#374151"),
        )
        cell_body_style = ParagraphStyle(
            "CellBody",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#1f2937"),
        )
        cell_body_right = ParagraphStyle(
            "CellBodyRight",
            parent=cell_body_style,
            alignment=2, # Right alignment
        )

        story = []

        # 1. Header
        period = report_data.get("period", {})
        summary = report_data.get("summary", {})
        breakdown = report_data.get("breakdown", [])

        gen_time = timezone.now().astimezone(IST_TZ).strftime("%d %b %Y, %I:%M %p IST")

        story.append(Paragraph("TOY STORE", subtitle_style))
        story.append(Spacer(1, 2))
        story.append(Paragraph("Executive Sales Report", title_style))
        story.append(Spacer(1, 6))

        meta_text = (
            f"<b>Report Period:</b> {period.get('start_date', '-')} to {period.get('end_date', '-')} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"<b>Grouping:</b> {str(period.get('group_by', 'day')).capitalize()} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"<b>Generated At:</b> {gen_time}"
        )
        story.append(Paragraph(meta_text, meta_style))
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e5e7eb"), spaceBefore=0, spaceAfter=12))

        # 2. Executive Summary Metrics Box
        story.append(Paragraph("Financial & Sales Summary", section_style))

        summary_table_data = [
            [
                Paragraph("<b>Net Sales</b>", cell_header_style),
                Paragraph(format_inr(summary.get("net_sales", 0)), cell_header_style),
                Paragraph("<b>Gross Sales</b>", cell_header_style),
                Paragraph(format_inr(summary.get("gross_sales", 0)), cell_header_style),
            ],
            [
                Paragraph("Total Orders", cell_body_style),
                Paragraph(str(summary.get("order_count", 0)), cell_body_style),
                Paragraph("Units Sold", cell_body_style),
                Paragraph(str(summary.get("units_sold", 0)), cell_body_style),
            ],
            [
                Paragraph("Offer Discounts", cell_body_style),
                Paragraph(format_inr(summary.get("offer_discount", 0)), cell_body_style),
                Paragraph("Coupon Discounts", cell_body_style),
                Paragraph(format_inr(summary.get("coupon_discount", 0)), cell_body_style),
            ],
            [
                Paragraph("Total Discounts", cell_body_style),
                Paragraph(format_inr(summary.get("total_discount", 0)), cell_body_style),
                Paragraph("Shipping Fees", cell_body_style),
                Paragraph(format_inr(summary.get("shipping", 0)), cell_body_style),
            ],
            [
                Paragraph("Cancelled Refunds", cell_body_style),
                Paragraph(format_inr(summary.get("cancelled_amount", 0)), cell_body_style),
                Paragraph("Returned Refunds", cell_body_style),
                Paragraph(format_inr(summary.get("returned_amount", 0)), cell_body_style),
            ],
        ]

        summary_table = Table(summary_table_data, colWidths=[120, 140, 120, 140])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f9fafb")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#f3f4f6")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#e0e7ff")), # Highlight net sales
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 14))

        # 3. Breakdown History Table
        story.append(Paragraph("Period Breakdown History", section_style))

        breakdown_table_data = [
            [
                Paragraph("<b>Date / Period</b>", cell_header_style),
                Paragraph("<b>Orders</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
                Paragraph("<b>Units</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
                Paragraph("<b>Gross</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
                Paragraph("<b>Discount</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
                Paragraph("<b>Cancelled</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
                Paragraph("<b>Returned</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
                Paragraph("<b>Net Sales</b>", ParagraphStyle("HR", parent=cell_header_style, alignment=2)),
            ]
        ]

        for row in breakdown:
            breakdown_table_data.append([
                Paragraph(str(row.get("date", "-")), cell_body_style),
                Paragraph(str(row.get("order_count", 0)), cell_body_right),
                Paragraph(str(row.get("units_sold", 0)), cell_body_right),
                Paragraph(format_inr(row.get("gross_sales", 0)), cell_body_right),
                Paragraph(format_inr(row.get("total_discount", 0)), cell_body_right),
                Paragraph(format_inr(row.get("cancelled_amount", 0)), cell_body_right),
                Paragraph(format_inr(row.get("returned_amount", 0)), cell_body_right),
                Paragraph(f"<b>{format_inr(row.get('net_sales', 0))}</b>", cell_body_right),
            ])

        breakdown_table = Table(breakdown_table_data, colWidths=[80, 45, 45, 75, 70, 65, 65, 78], repeatRows=1)
        breakdown_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(breakdown_table)

        doc.build(story, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        return buffer


class SalesReportExcelGenerator:

    @classmethod
    def generate(cls, report_data):
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(name="Calibri", size=14, bold=True, color="1E1B4B")
        sub_font = Font(name="Calibri", size=10, italic=True, color="4B5563")
        tbl_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)

        tbl_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        summary_header_fill = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")
        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        currency_fmt = "₹#,##0.00"
        number_fmt = "#,##0"

        period = report_data.get("period", {})
        summary = report_data.get("summary", {})
        breakdown = report_data.get("breakdown", [])
        gen_time = timezone.now().astimezone(IST_TZ).strftime("%Y-%m-%d %H:%M:%S IST")

        # ---------------------------------------------------------
        # Sheet 1: Summary
        # ---------------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Summary"

        ws_sum.append(["TOY STORE - Sales Report Summary"])
        ws_sum.cell(row=1, column=1).font = header_font
        ws_sum.append([f"Report Period: {period.get('start_date')} to {period.get('end_date')} | Grouping: {period.get('group_by')} | Generated: {gen_time}"])
        ws_sum.cell(row=2, column=1).font = sub_font
        ws_sum.append([])

        # Table Header
        ws_sum.append(["Metric", "Value"])
        ws_sum.cell(row=4, column=1).font = tbl_header_font
        ws_sum.cell(row=4, column=1).fill = summary_header_fill
        ws_sum.cell(row=4, column=2).font = tbl_header_font
        ws_sum.cell(row=4, column=2).fill = summary_header_fill

        metrics_rows = [
            ("Total Orders", summary.get("order_count", 0), number_fmt),
            ("Units Sold", summary.get("units_sold", 0), number_fmt),
            ("Gross Sales", float(summary.get("gross_sales", 0)), currency_fmt),
            ("Offer Discounts", float(summary.get("offer_discount", 0)), currency_fmt),
            ("Coupon Discounts", float(summary.get("coupon_discount", 0)), currency_fmt),
            ("Total Discounts", float(summary.get("total_discount", 0)), currency_fmt),
            ("Shipping Fees", float(summary.get("shipping", 0)), currency_fmt),
            ("Cancelled Refunds", float(summary.get("cancelled_amount", 0)), currency_fmt),
            ("Returned Refunds", float(summary.get("returned_amount", 0)), currency_fmt),
            ("Refunded Amount (Total)", float(summary.get("refunded_amount", 0)), currency_fmt),
            ("Net Sales", float(summary.get("net_sales", 0)), currency_fmt),
        ]

        for idx, (label, val, fmt) in enumerate(metrics_rows, start=5):
            ws_sum.append([label, val])
            c_label = ws_sum.cell(row=idx, column=1)
            c_val = ws_sum.cell(row=idx, column=2)

            c_label.font = bold_font if label == "Net Sales" else regular_font
            c_val.font = bold_font if label == "Net Sales" else regular_font
            c_val.number_format = fmt
            c_label.border = thin_border
            c_val.border = thin_border
            if idx % 2 == 0:
                c_label.fill = zebra_fill
                c_val.fill = zebra_fill

        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 22

        # ---------------------------------------------------------
        # Sheet 2: Sales Breakdown
        # ---------------------------------------------------------
        ws_bd = wb.create_sheet(title="Sales Breakdown")

        ws_bd.append(["TOY STORE - Sales Breakdown History"])
        ws_bd.cell(row=1, column=1).font = header_font
        ws_bd.append([f"Period: {period.get('start_date')} to {period.get('end_date')} | Grouping: {period.get('group_by')}"])
        ws_bd.cell(row=2, column=1).font = sub_font
        ws_bd.append([])

        bd_headers = [
            "Period / Date", "Orders", "Units Sold", "Gross Sales",
            "Offer Discount", "Coupon Discount", "Total Discount",
            "Shipping", "Cancelled Amount", "Returned Amount", "Net Sales"
        ]
        ws_bd.append(bd_headers)
        header_row_idx = 4

        for col_idx in range(1, len(bd_headers) + 1):
            cell = ws_bd.cell(row=header_row_idx, column=col_idx)
            cell.font = tbl_header_font
            cell.fill = tbl_header_fill
            cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "right", vertical="center")

        ws_bd.freeze_panes = "A5"

        for row_idx, row in enumerate(breakdown, start=5):
            bd_values = [
                str(row.get("date", "")),
                int(row.get("order_count", 0)),
                int(row.get("units_sold", 0)),
                float(row.get("gross_sales", 0)),
                float(row.get("offer_discount", 0)),
                float(row.get("coupon_discount", 0)),
                float(row.get("total_discount", 0)),
                float(row.get("shipping", 0)),
                float(row.get("cancelled_amount", 0)),
                float(row.get("returned_amount", 0)),
                float(row.get("net_sales", 0)),
            ]
            ws_bd.append(bd_values)

            for col_idx in range(1, len(bd_values) + 1):
                cell = ws_bd.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in (2, 3):
                    cell.number_format = number_fmt
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.number_format = currency_fmt
                    cell.alignment = Alignment(horizontal="right")

                if row_idx % 2 == 1:
                    cell.fill = zebra_fill

        # Auto column widths
        for col in ws_bd.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_bd.column_dimensions[col_letter].width = max(max_len + 4, 14)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
